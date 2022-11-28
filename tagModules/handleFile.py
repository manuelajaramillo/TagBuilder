"""Wrapper functions to handle TR File.
This module integrate some functionalities from Openpyxl in xlsx class
that allow us to manipulate and handle .xlsx format files as TR file that 
is used in GroupM Nexus in the process of implement a measument strategy.
"""
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from datetime import date as dt

from os import path as p
import re

MONTHS  = [
    'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
    'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
    ]

TRIGGER = [
    'PV', 'Click', 'Scroll'
    ]

SHEET_NAME = 'Tagging Request'

PATH_      = 'resources/formats/TaggingRequest_2021.xlsx'

FONT = {
    'name':'Calibri', 'size':11, 'bold':False, 'italic':False, 'color':'FF000000'
    }

LOGO_LOC = 'resources/Nexus_Logo.png'

class xlsxFile:
    def __init__(self, sheet = 'Tagging Request', PATH=PATH_):
        """Constructor Method
        Three important processes are carried out in this constructor method, namely:
            1. Access to TR file base. 
            2. Initialize a workbook as attribute from TR file base.
            3. Set up the sheet Tagging Request as work and active sheet.

        Args:
            sheet (str, optional): The base sheet name that contains the main format in TR file. Defaults to 'Tagging Request'.
            PATH (str, optional): Route to access to TR file base require to implement a Measurement Strategy. Defaults to PATH_.
        """        
        self.PATH = PATH
        self.setBook()
        self.sheet = self.book[sheet]
        
    def setPATH(self, path=PATH_):
        """This method sets the path to the TR file as class attribute.

        Args:
            path (str, optional): Path to the TR file. Defaults to path to the TR file base.
        """        
        self.PATH = path
    
    def setBook(self, path=None):
        """This method accesses and retrieves the .xlsx file that the PATH attribute
        specify and set and initialize the workbook attribute of the class.

        Args:
            path (str, optional): Path to the TR file. Defaults to None.
        """               
        if path != None: self.setPATH(path)
        self.book = load_workbook(p.abspath(self.PATH)) 
        
    def saveBook(self, dir_path = None, default_name=True, bookName=None):
        """This method generates a TR file named as TagReq_AdvertiserName_MonthYear.xlsx 
            Month Format: First three letters of the month in english as July -> Jul,
                April -> Apr
            Year Format: For number format as 2017.

        Args:
            dir_path (str, optional)     : Path directory to save the TR File. Defaults to None.
            default_name (bool, optional): Boolean variable to determine if it is require so assign a default name to the file. Defaults to True.
            bookName (bool, optional)    : Name to assign to the file. Defaults to True.

        Returns:
            str: Path where the TR file was saved.
        """        
        if default_name:
            nameFile = self.getFileName('C13')
        else:
            if bookName != None:
                nameFile = bookName
            else:
                nameFile = self.getFileName('C13', 'Final')
        if dir_path == None:
            self.book.save(nameFile)
            return p.abspath(nameFile).replace('\\','/')
        else:
            self.book.save(p.join(dir_path, nameFile))
            return dir_path+'/'+nameFile
        
    def setSheet(self, nameSheet = SHEET_NAME):
        """This method sets the attribute sheet to the the sheet name given.

        Args:
            nameSheet (str, optional): Sheet name given to set the sheet. Defaults to Tagging Request Sheet.
        """        
        self.sheet = self.book[nameSheet]
        
    def duplicateSheet(self, nameSheetFrom, nameSheetTo):
        """This method duplicates a sheet given.

        Args:
            nameSheetFrom (str) : Name of the sheet to be duplicated.
            nameSheetTo (str)   : Name of the sheet duplicated.
        """        
        self.setSheet(nameSheetFrom)
        self.book.copy_worksheet(self.sheet).title = nameSheetTo
        
    def insertImage(self, sheetName, cell):
        self.setSheet(sheetName)
        self.sheet.add_image(openpyxl.drawing.image.Image(LOGO_LOC), cell)
        
    def setSectionSheets(self, nameSheetFrom, sectionList):
        """This methods create a new sheets from a sheet given.

        Args:
            nameSheetFrom (str): Name of the sheet base to be duplicated.
            sectionList (list or tuple): Array list with the names of the sheet to be created.
        """        
        for section in sectionList:
            self.duplicateSheet(nameSheetFrom, section)
    
    def loadList(self, dataList, cellOrigin = 'F31'):
        """This method allow us to bulk a data to the active sheet in the workbook.

        Args:
            dataList (list): Array list with the data to write in the sheet.
            cellOrigin (str, optional): Origin cell from to write the data. Defaults to 'F31'.
        """        
        cell = cellOrigin
        for item in dataList:
            cell = self.getCellDown(cell)
            self.sheet[cell] = item

    def readCell(self, cell):
        """This method retrieves the information contains in a given cell.

        Args:
            cell (str): Name of the cell to read. For example: C31.

        Returns:
            str: The data retrieves from a cell given.
        """        
        return self.sheet[cell].value
    
    def readNextCell(self, cell, direction='vertical'):
        """Function to get the value store in the neighborhood cells.

        Args:
            cell (str): Name cell with excel nomenclature. e.g E32
            direction (str, optional): Direction of the search of our cell neighborhood. Defaults to 'vertical'.

        Returns:
            Value Cell: The value stores in the neighborhood cell.
        """
        cell = cell.upper()
        if direction == 'vertical':
            nextCell = re.findall(r'\D+', cell)[0]+str(int(re.findall(r'\d+', cell)[0])+1)
            return nextCell, self.readCell(nextCell)
        else:
            row, col = int(re.findall(r'\d+', cell)[0]), self.getColIndex(re.findall(r'\D+', cell)[0])+1
            return self.sheet.cell(row, col).coordinate, self.sheet.cell(row, col).value
            
    def getColIndex(self, colString):
        """Function to mappear from name column domain to numeric column index.

        Args:
            colString (str): Name of the column in excel format: AAK, A, etc

        Returns:
            colIndex: column index.
        """
        colIndex = 0
        for c,e in zip(colString, range(len(colString)-1,-1,-1)):
            colIndex = colIndex + (ord(c)-64)*pow(26, e)
        return colIndex
    
    def writeCell(self, cell, value, aligment_=['center','center'], font_=FONT):
        """This method write in a cell given the data passed in value parameter.

        Args:
            cell (str): Name of the cell to write. For example: C31.
            value (str): The data to be written.
            aligment_ (list, optional): Array list with the aligment style of the cell. Defaults to ['center','center'].
        """        
        self.sheet[cell] = value
        cell = self.sheet[cell]
        cell.font = Font(name=font_['name'], size=font_['size'], bold=font_['bold'], italic=font_['italic'], color=font_['color'])
        cell.alignment = Alignment(horizontal=aligment_[0], vertical=aligment_[1], wrapText=True)

    def getCellDown(self, cell):
        """This method returns the name of the cell in the position immediately below.

        Args:
            cell (str): Name of the cell reference.

        Returns:
            str: A name cell in the format LetterColumn:NumberRow. For example: A31.
        """        
        for i in range(0,len(cell)):
            if cell[i].isdigit():
                return cell[0:i]+str(int(cell[i:])+1)
            
    def getCellUp(self, cell):
        """This method returns the name of the cell in the position immediately above.

        Args:
            cell (str): Name of the cell reference.

        Returns:
            str: A name cell in the format LetterColumn:NumberRow. For example: A31.
        """        
        for i in range(0,len(cell)):
            if cell[i].isdigit():
                return cell[0:i]+str(int(cell[i:])-1)
            
    def nextFreeCell(self, cell, rowCell=True):
        if rowCell:
            column = re.findall(r'\D+',cell)[0]
            row    = int(re.findall(r'\d+',cell)[0])
            while True:
                if self.readCell(cell) in [None, '']:
                    return cell, row, column
                row += 1
                cell = column+str(row)
        else:
            #We need to implement horizontal searching
            return cell, None, None
            
    def getLastPath(self, path):
        """This method retrieve the last subpath in a given path.

        Args:
            path (str): Path to extract the subpath.

        Returns:
            str: Subpath extracted.
        """        
        return path[1:].split('/')[0]
            
    def getDate(self):
        """This method retrieves the current date.

        Returns:
            tuple: MonthYear format. For example: Apr2022.
        """        
        month = ''
        for i in range(1,13):
            if dt.today().month == i:
                month = MONTHS[i-1]
        return month, str(dt.today().year)
            
    def getFileName(self, cell, custom_text=''):
        """This method creates and retuns the name to be assigned to TR File.

        Args:
            cell (str)                  : Name of cell with the name of the advertiser.
            custom_text (str, optional) : Aditional word text to be added to the name of TR file. Defaults to ''.

        Returns:
            str: The name to be assigned to the TR file.
        """        
        advertiser  = self.readCell(cell)
        month, year = self.getDate()
        if advertiser != None:
            return 'TagReq'+custom_text+'_'+advertiser+'_'+month+year+'.xlsx'
        else:
            return 'TagReq'+custom_text+'_'+'advertiserNameNoFound'+'_'+month+year+'.xlsx'

    def getNameSection(self, advertiserName, sectionName, typeTrigger='PV'):
        """This method create and return the standar name of a pixel to create in TagBuilder.

        Args:
            advertiserName (str): Advertiser Name
            sectionName (str): Section Name.
            typeTrigger (str, optional): Type of trigger. Defaults to 'PV'.

        Returns:
            str: The name of a pixel to create in TagBuilder.
        """        
        month, year = self.getDate()
        return advertiserName+'_'+sectionName+typeTrigger+'_'+month+year
                  
    def getTagName(self, path, cell):
        """_summary_

        Args:
            path (_type_): _description_
            cell (_type_): _description_

        Returns:
            _type_: _description_
        """        
        advertiser  = self.readCell(cell) 
        month, year = self.getDate()
        lastPath    = self.getLastPath(path)
        trigger     = 'PV'
        if lastPath == '':
            lastPath = 'Home'
        return advertiser+'_'+lastPath.capitalize()+trigger+'_'+month+year+'.xlsx'
    
    def fillCell(self, cell, color):
        thin = Side(border_style="thin", color="FFFFFF")
        self.sheet[cell].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        self.sheet[cell].fill = PatternFill("solid", fgColor=color)
        
            
if __name__ == '__main__':
    file = xlsxFile()
